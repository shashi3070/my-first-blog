
import xlrd, xlwt
from pytz import timezone
from xlutils.copy import copy as xl_copy
import pyodbc
import os
import psycopg2
from boto.s3.connection import S3Connection
from boto.s3.key import Key
from datetime import datetime as dt
from openpyxl import load_workbook
import logging
from logging.handlers import RotatingFileHandler
import openpyxl
from openpyxl.styles.borders import Border, Side, BORDER_THIN
logger=''
#from Helpers.Mails import SendMailWithoutAttachment
#from Mails_With_Attachment import SendMailWithAttchment
#from Helper.Mails_With_Attachment import SendMailWithAttchment
OutputPath='D:/Outbound_Zip/Input_output/'
filePath='D:/Outbound_Zip/Input_output/'
LogoPath1='D:/Outbound_Zip/Input_output/header_sc_mat1.png'
LogoPath2='D:/Outbound_Zip/Input_output/header_sc_max1.png'
Header_Query={}
Summary_Query={}
Summary_Loc_Map={}
Header_Loc_Map={}


filename=''
OutputFileName=''
SheetName=''
subject=''
EmailBodyContent=''
frommail=''
ReceiverList=''
password=''
custom_file_name=''

def print_sql_table(table):
	for row in table:
		print(row)

def SqlConn():
	try:
		conn=pyodbc.connect('driver={%s};server=%s;database=%s;uid=%s;pwd=%s' % ('SQL Server', '172.31.23.171', 'salesiq_IC_processing_DEV', 'salesiq_IC_processing_dev', 'salesiq@123' ) )
		print('Sql connection successful')
		logger.info('Sql connection successful')
		return conn
	except Exception as E:
		logger.error('in SqlConn Error '+str(E))
		return 'in SqlConn Error '+str(E)

def InitGlobalVar(type):
	logger.info('-----------------InitGlobalVar------------------')
	global subject,EmailBodyContent,frommail,ReceiverList,OutputFileName,SheetName,filename,password,custom_file_name
	conn=SqlConn()
	query='select [subject],message_html,[From],recipient,[File_Name],Report_SheetName,Sample_FileName_Report,EMail_Pass from [dbo].[AXTRIA_SYSTEM_NOTIFICATION] where email_template=?'
	cursor = conn.cursor()
	table=cursor.execute(query,type)
	#print table
	#print_sql_table(table)
	for info in table:
		subject=info[0]
		EmailBodyContent=info[1]
		frommail=info[2]
		ReceiverList=info[3]
		ReceiverList=ReceiverList.split(',')
		OutputFileName=info[4]
		custom_file_name=OutputFileName
		SheetName=info[5]
		filename=info[6]
		password=info[7]
	print('-----------------printing Global varibale valaues--------------------')
	logger.info('-----------------printing Global varibale valaues--------------------')

	print('subject----'+str(subject))
	logger.info('subject----'+str(subject))
	print('EmailBodyContent----'+str(EmailBodyContent))
	logger.info('EmailBodyContent----'+str(EmailBodyContent))
	print('ReceiverList----'+str(ReceiverList))
	logger.info('ReceiverList----'+str(ReceiverList))
	print('OutputFileName----'+str(OutputFileName))
	logger.info('OutputFileName----'+str(OutputFileName))
	print('custom_file_name----'+str(custom_file_name))
	logger.info('custom_file_name----'+str(custom_file_name))
	print('SheetName----'+str(SheetName))
	logger.info('SheetName----'+str(SheetName))
	print('filename----'+str(filename))
	logger.info('filename----'+str(filename))
	print('password----'+str(password))
	logger.info('password----'+str(password))

	cursor.close()
	conn.close()

def init():
	logger.info('start----init')
	Summary_Query[6]='''select * from JNJ_DB.VAL_SC_MOD3_DATA_COMP_FLAG'''
	Summary_Query[5]='''select * from JNJ_DB.VAL_SC_MOD2_COMP_FLAG'''
	Summary_Query[4]='''select * from JNJ_DB.VAL_SC_MOD1_DATA_COMP_FLAG'''
	Summary_Query[3]='''select * from JNJ_DB.val_s_growth_objectives_comp_flag'''
	Summary_Query[2]='''select * from jnj_db.val_base_objectives_data_comp_flag'''
	Summary_Query[1]='''select * from JNJ_DB.val_pas_comp_flag'''
	Summary_Loc_Map[1]={'row':10,'col':3}
	Summary_Loc_Map[2]={'row':10,'col':8}
	Summary_Loc_Map[3]={'row':10,'col':16}
	Summary_Loc_Map[4]={'row':10,'col':24}
	Summary_Loc_Map[5]={'row':10,'col':32}
	Summary_Loc_Map[6]={'row':10,'col':40}
	Header_Query[1]='''Select '# of Records in SIQ' as SYS, count(*) from jnj_db.PAS_SIQ UNION Select '# of Records in MATILLION'  , count(*) from jnj_db.M_PAS_MASTER '''
	Header_Query[2]='''Select '# of Records in SIQ' as SYS, count(*) from jnj_db.S_Base_Objectives_SIQ UNION  Select '# of Records in MATILLION'  , count(*) from jnj_db.M_Q_MOD_2_Base_Objective ;'''
	Header_Query[3]='''Select '# of Records in SIQ' as SYS, count(*) from jnj_db.S_Forecast_Objectives_SIQ UNION  Select '# of Records in MATILLION'  , count(*) from jnj_db.M_Q_MOD_2_Growth_Adjust;'''
	Header_Query[4]='''Select '# of Records in SIQ' as SYS, count(*) from jnj_db.m_sc_mod01_siq UNION  Select '# of Records in MATILLION'  , count(*) from jnj_db.M_SC_MOD_1_3_A'''
	Header_Query[5]='''Select '# of Records in SIQ' as SYS, count(*) from jnj_db.SC_MOD2_SIQ UNION  Select '# of Records in MATILLION'  , count(*) from jnj_db.M_SC_MOD_2_2_A;'''
	Header_Query[6]='''Select '# of Records in SIQ' as SYS, count(*) from jnj_db.SC_MOD3_SIQ UNION  Select '# of Records in MATILLION'  , count(*) from jnj_db.M_SC_MOD_3_1_A;'''
	Header_Loc_Map[1]={'row':5,'col':3}
	Header_Loc_Map[2]={'row':5,'col':8}
	Header_Loc_Map[3]={'row':5,'col':16}
	Header_Loc_Map[4]={'row':5,'col':24}
	Header_Loc_Map[5]={'row':5,'col':32}
	Header_Loc_Map[6]={'row':5,'col':40}
	logger.info('end----init')

def OpenSheet(SheetName,filePath,filename):
    logger.info('OpenSheet start')
    print('--------------')
    wb = load_workbook(filePath+filename)
    ws = wb[SheetName]
    return (ws,wb)
    logger.info('OpenSheet end')
    
def FetchSummaryData(query):
	logger.info('FetchSummaryData start')
	try:
		print('Ín FetchSummaryData')
		redshift_con =psycopg2.connect(dbname= 'jandj', host='j-jmedical-devices-2195.coya9jnvir3f.us-west-2.redshift.amazonaws.com', port= '5439', user= 'admin', password= 'Welcome123')
		#print("Executing " + query)
		cur = redshift_con.cursor()
		cur.execute(query)
		A=cur.fetchall()
		#print(A)
		cur.close()
		redshift_con.close()
		logger.info('Table FetchSummaryData  '+str(A))
		return A
	except Exception as Ex:
		logger.error('Error in redshift connection' + str(Ex))
		return 'Error in redshift connection' + str(Ex)

def WriteSummaryTables(temp_table,row,col,sheetvar,wb):
	logger.info('start WriteSummaryTables')

	row=row-1
	col_init=col-1
	first_col=col
	for i in temp_table:
		row=row+1
		col=col_init
		for j in i:
			print(j)
			col=col+1
			if first_col==col:
				c = sheetvar.cell(row , col)
				c.value = str(j).upper()
			else:
				c = sheetvar.cell(row , col)
				c.value = j
	logger.info('end WriteSummaryTables')

def Delinputfile(file):
	logger.info('start Delinputfile')
	if os.path.exists(OutputPath+file):
		os.remove(OutputPath+file)
		print('File Deleted')
		logger.info('File Deleted'+str(file))


def AddHeaderData(ws,wb):
	logger.info('start AddHeaderData')
	for index in Header_Query:
		query=Header_Query[index]
		print('AddSummaryData--query---'+str(query))
		table=FetchSummaryData(query)
		loc=Header_Loc_Map[index]
		print(loc)
		row=loc['row']
		col=loc['col']
		print('Details of tables filling----')
		print('---Index, row,col----'+str(index)+'---'+str(row)+"-----"+str(col))
		WriteSummaryTables(table,row,col,ws,wb)
	logger.info('end AddHeaderData')

def AddSummaryData(ws,wb):
	logger.info('start AddSummaryData')
	for index in Summary_Query:
		query=Summary_Query[index]
		print('AddSummaryData--query---'+str(query))
		table=FetchSummaryData(query)
		loc=Summary_Loc_Map[index]
		print(loc)
		row=loc['row']
		col=loc['col']
		print('Details of tables filling----')
		logger.info('Details of tables filling----')
		print('---Index, row,col----'+str(index)+'---'+str(row)+"-----"+str(col))
		WriteSummaryTables(table,row,col,ws,wb)
		logger.info('end  AddSummaryData')

def test():
	print('Test')

def addHeaderImage(ws,wb):
	logger.info('start  addHeaderImage')
	img =openpyxl.drawing.image.Image(LogoPath1)
	ws.add_image(img,'Q1')
	img =openpyxl.drawing.image.Image(LogoPath2)
	ws.add_image(img,'J1')
	logger.info('end  addHeaderImage')

def HideSheet(sheetname,wb):
	sheet_ref=wb[sheetname]
	sheet_ref.sheet_state='hidden'

def MasterDef():
	logger.info('start  MasterDef')
	wb=OpenSheet(SheetName,filePath,filename)
	ws=wb[0]
	wb=wb[1]
	addHeaderImage(ws,wb)
	AddSummaryData(ws,wb)
	AddHeaderData(ws,wb)
	HideSheet(SheetName,wb)

	logger.info('end  MasterDef')



	wb.save(OutputPath+OutputFileName)
	wb.close()

def SC_Variance_Report():
	try:
		global logger
		handlers = [RotatingFileHandler(filename=OutputPath+str(dt.now().strftime('%m-%d-%Y'))+'logs.txt', mode='a', maxBytes=512000,backupCount=4)]
		logging.basicConfig(handlers=handlers, level=logging.DEBUG, format='%(levelname)s %(asctime)s %(message)s', datefmt='%m/%d/%Y%I:%M:%S %p')
		logger = logging.getLogger('my_logger')
		logger.info('--------------Process start--------------------------')
		
		InitGlobalVar('Sales_Crediting_Data_Variance_Report')
		init()
		Delinputfile(OutputFileName)
		MasterDef()
		return 'Success'
	except Exception as E:
		
		return str(E)
	#SendMailWithAttchment(frommail,subject,OutputPath,OutputFileName,password,ReceiverList,custom_file_name,EmailBodyContent)

SC_Variance_Report()


